#cd C:\Users\Nitish\Desktop\volume_scanner
#env\Scripts\Activate

import time
import json
import os
from datetime import datetime
import requests
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from logzero import logger
from SmartApi.smartConnect import SmartConnect
import pyotp

# --- Credentials ---
API_KEY = "X1q82Ysv"
CLIENT_CODE = "AAAO662742"
MPIN = "3010"
TOTP_SECRET = "IWXYJI7S26EUXGNWB4OCNHUQ7A"

# --- Telegram Credentials ---
TELEGRAM_TOKEN = "7632650061:AAFWRVJuIKCOwsuzio9j3HVDp5xWq4ki4gA"
TELEGRAM_CHAT_ID = "6712530011"

# --- Send Telegram Message ---
def send_telegram_message(message):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    payload = {
        "chat_id": TELEGRAM_CHAT_ID,
        "text": message,
        "parse_mode": "Markdown"
    }
    try:
        response = requests.post(url, data=payload)
        print("📬 Telegram HTTP status:", response.status_code)
        print("📬 Telegram response body:", response.text)
        if response.status_code != 200:
            print("❌ Telegram send failed.")
        else:
            print("✅ Telegram message accepted by API.")
    except Exception as e:
        print("❌ Exception while sending Telegram message:", e)

# --- Load NSE EQ symbols ---
with open("verified_nse_eq.json", "r") as f:
    all_stocks = json.load(f)
    HARDCODED_EQUITIES = [s for s in all_stocks if s.get("symbol", "").endswith("-EQ") and s.get("token")]

# --- SmartAPI Login ---
smart_api = SmartConnect(api_key=API_KEY)
totp = pyotp.TOTP(TOTP_SECRET).now()

try:
    data = smart_api.generateSession(clientCode=CLIENT_CODE, password=MPIN, totp=totp)
    feed_token = smart_api.getfeedToken()
    logger.info("✅ Login successful via SmartAPI.")
except Exception as e:
    logger.error(f"❌ Login failed: {e}")
    send_telegram_message("❌ SmartAPI login failed.")
    exit()

# --- Candle Data Fetch ---
def get_candle_data(symbol, token, interval="ONE_DAY", max_retries=3, backoff=2):
    for attempt in range(1, max_retries + 1):
        try:
            params = {
                "exchange": "NSE",
                "symboltoken": token,
                "interval": interval,
                "fromdate": "2023-06-01 09:15",
                "todate": time.strftime("%Y-%m-%d 15:30")
            }
            response = smart_api.getCandleData(params)
            if "data" in response and response["data"]:
                return response["data"]
            return None
        except Exception as e:
            logger.warning(f"[{symbol}] Attempt {attempt}: {e}")
            if attempt == 2:
                try:
                    logger.info("🔁 Re-logging SmartAPI...")
                    totp = pyotp.TOTP(TOTP_SECRET).now()
                    smart_api.terminateSession(CLIENT_CODE)
                    smart_api.generateSession(clientCode=CLIENT_CODE, password=MPIN, totp=totp)
                except Exception as inner:
                    logger.error(f"❌ Re-login failed: {inner}")
            time.sleep(backoff)
    return None

# --- EMA Calculation ---
def calculate_emas(df, column, prefix):
    periods = {
        f"{prefix}EMA_5": 5,
        f"{prefix}EMA_10": 10,
        f"{prefix}EMA_20": 20,
        f"{prefix}EMA_50": 50,
        f"{prefix}EMA_100": 100,
        f"{prefix}EMA_200": 200
    }
    for name, period in periods.items():
        df[name] = df[column].ewm(span=period, adjust=False).mean()
    return df

# --- Main Script ---
def main():
    summary_rows = []
    alert_rows = []

    for stock in HARDCODED_EQUITIES:
        symbol = stock["symbol"]
        token = stock["token"]
        logger.info(f"📊 Processing {symbol}...")
        candle_data = get_candle_data(symbol, token)
        if not candle_data:
            continue

        df = pd.DataFrame(candle_data, columns=["Datetime", "Open", "High", "Low", "Close", "Volume"])
        df["Datetime"] = pd.to_datetime(df["Datetime"])
        df.sort_values("Datetime", inplace=True)

        df = calculate_emas(df, "Volume", "Vol_")
        df = calculate_emas(df, "Close", "Close_")

        latest = df.iloc[-1]

        row = {
            "Symbol": symbol,
            "Date": latest["Datetime"].date(),
            "Open": latest["Open"],
            "High": latest["High"],
            "Low": latest["Low"],
            "Close": latest["Close"],
            "Volume": latest["Volume"]
        }

        highlight = False
        volx5 = None
        periods_to_check = [5, 10, 20, 50, 100, 200]

        for p in periods_to_check:
            vol_ema = latest.get(f"Vol_EMA_{p}")
            volx = round(latest["Volume"] / vol_ema, 2) if vol_ema else None
            row[f"Volx{p}"] = volx
            row[f"EMA{p}_Vol"] = vol_ema
            row[f"EMA{p}_Close"] = latest.get(f"Close_EMA_{p}")

            if p == 5:
                volx5 = volx
            if volx and volx >= 2:
                highlight = True

        summary_rows.append(row)
        if highlight:
            alert_rows.append(row)
        time.sleep(1)

    df_summary = pd.DataFrame(summary_rows)
    sheet_name = datetime.now().strftime("%Y-%m-%d")
    excel_file = "nifty_volume_summary.xlsx"

    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)

    for col_idx, col in enumerate(df_summary.columns, 1):
        ws.cell(row=1, column=col_idx, value=col)
        for row_idx, val in enumerate(df_summary[col], 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col == "Volx5":
                if val is not None:
                    if val >= 10:
                        fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")  # Purple
                        cell.fill = fill
                    elif val >= 5:
                        fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
                        cell.fill = fill
                    elif val >= 2:
                        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
                        cell.fill = fill

    wb.save(excel_file)
    logger.info(f"✅ Excel saved: {sheet_name}")

    MAX_ROWS = 50
    if alert_rows:
        msg_lines = [
            f"📊 *Volume Spike Summary* ({sheet_name})",
            "```",
            f"{'Symbol':<12} {'Close':>8} {'Volx5':>7} {'Volx10':>7} {'Volx20':>7}",
            f"{'-'*12} {'-'*8} {'-'*7} {'-'*7} {'-'*7}"
        ]

        for r in alert_rows[:MAX_ROWS]:
            emoji = ""
            v = r.get("Volx5")
            if v is not None:
                if v >= 10:
                    emoji = "🟣"
                elif v >= 5:
                    emoji = "🟢"
                elif v >= 2:
                    emoji = "🟡"

            msg_lines.append(
                f"{emoji} {r['Symbol']:<10} {r['Close']:>8.2f} {r.get('Volx5') or '':>7} {r.get('Volx10') or '':>7} {r.get('Volx20') or '':>7}"
            )

        msg_lines.append("```")
        message = "\n".join(msg_lines)
        send_telegram_message(message)
    else:
        send_telegram_message(f"📊 No significant volume spikes on {sheet_name}.")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.exception("❌ Script crashed.")
        send_telegram_message(f"❌ Script crashed: {e}")
